import cv2
import os
from PIL import Image
import openpyxl
from shutil import copy2


def add_filenames_to_excel(input_path, output_path, excel_path):
    file = openpyxl.load_workbook(excel_path)#"auction_qr_sample.xlsx")
    current_sheet = file[file.sheetnames[0]]
    #current_sheet = file["AUCTION QR SAMPLE(11672)"]
    print(current_sheet["B2"].value)
    columns_to_add = "IJKLMF"

    barcode_image_dict = pick_up_images(input_path)
    barcode_filename_dict = rename_and_sort_images(barcode_image_dict,input_path=input_path, output_path=output_path)

    for row in range(2, current_sheet.max_row + 1):
        print(row)
        column = "H"
        cell_name = f"{column}{row}"
        print(cell_name)
        print(current_sheet[cell_name].value)
        sheet_barcode = current_sheet[cell_name].value

        for code_img in barcode_filename_dict:
            if str(sheet_barcode) == code_img:
                print(f"adding to {code_img} these images {barcode_filename_dict[code_img]}")
                column_counter = 0

                for img_filename in barcode_filename_dict[code_img]:
                    # for x in range(0, len(columns_to_add)):
                    col = columns_to_add[column_counter]

                    #render_image = openpyxl.drawing.image.Image(fr"images/{img}")
                    #render_image.height = 50
                    #render_image.width = 50
                    try:
                        current_sheet[f"{col}{row}"] = img_filename
                        #render_image.anchor = f"{col}{row}"
                        # current_sheet[f"{col}{row}"].value = render_image
                        #current_sheet.add_image(render_image)
                        print(f"added to {col}{row} image: {img_filename}")
                        column_counter += 1
                    except:
                        pass
    excel_output = excel_path.split(".xlsx")

    # Save the processed excel file to same path of original excel file
    file.save(f'{excel_output[0]}_complete.xlsx')

    return None


def rename_and_sort_images(barcode_image_dict, input_path, output_path):
    # takes image dict and input/output path
    # places new renamed images in output path

    barcode_filename_dict = {}
    image_group_counter = 1
    for barcode in barcode_image_dict:
        print("BARCODE")
        print(barcode)
        image_individual_counter = 1

        sorted_filename_list = []

        for image_filename in barcode_image_dict[barcode]:
            print(image_filename)

            sorted_filename = f"{image_group_counter}-{image_individual_counter}.jpg"

            img_path = copy2(f"{input_path}/{image_filename}",
                             f"{output_path}/{sorted_filename}")

            image_individual_counter += 1

            sorted_filename_list.append(sorted_filename)
            print(img_path)

        barcode_filename_dict[barcode] = sorted_filename_list
        image_group_counter += 1

    print(f"barcode file name dict: {barcode_filename_dict}")
    return barcode_filename_dict


def pick_up_images(image_path):
    path = image_path #r"images"
    image_list = os.listdir(path)
    print(f"image list: {image_list}")

    barcode_image_dict = {}
    prev_barcode = ""

    for image_file in image_list:
        qr_text = decode_qr(image_file, image_path)
        if qr_text:
            print(f"{image_file}")
            print(f"QR Value: {qr_text}")
            prev_barcode = qr_text
            barcode_image_dict[qr_text] = []
        if not qr_text:
            barcode_image_dict[prev_barcode].append(image_file)

    print(f"Barcode and images: {barcode_image_dict}")
    return barcode_image_dict


def decode_qr(image_filename, image_path):
    img = cv2.imread(f"{image_path}/{image_filename}", flags=cv2.IMREAD_REDUCED_GRAYSCALE_8)  # flags=cv2.IMREAD_GRAYSCALE)
    # cv2.imshow('image',img)
    # cv2.waitKey(0)
    # img2 = Image.open(r"images/DSC04890.jpg")
    # width, height = Image.open(r"images/DSC04890.jpg").size

    # print(width*height)
    # img2.show()
    detector = cv2.QRCodeDetector()

    val, pts, st_code = detector.detectAndDecode(img)

    # print(f"value: {val}")
    # print(f"Coordinates: {pts}")
    # print(f"{st_code}")

    if val:
        return val

# pick_up_images()

# add_images_to_file()

# decode_qr(f"DSC04898.jpg")
